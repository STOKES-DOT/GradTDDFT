from __future__ import annotations

import argparse
import csv
import gc
import json
import os
from dataclasses import dataclass, replace
from pathlib import Path
import sys
import time
from typing import Any

os.environ.setdefault("XLA_PYTHON_CLIENT_PREALLOCATE", "false")
os.environ.setdefault("XLA_PYTHON_CLIENT_ALLOCATOR", "platform")
os.environ.setdefault("MPLCONFIGDIR", str(Path("outputs") / ".mplconfig"))

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

import h5py
import jax

jax.config.update("jax_enable_x64", True)

import jax.numpy as jnp
import numpy as np
import optax
from pyscf import cc, dft, gto, scf

from td_graddft import neural_xc
from td_graddft.data.hdf5_cache import read_restricted_molecule, write_restricted_molecule
from td_graddft.data.reference import restricted_reference_from_pyscf
from td_graddft.neural_xc import (
    DEFAULT_INPUT_FEATURE_MODE,
    DEFAULT_NETWORK_ARCHITECTURE,
    DEFAULT_NETWORK_HIDDEN_DIMS,
)
from td_graddft.training import (
    GroundStateCoreDatum,
    GroundStateCoreTrainingConfig,
    GroundStateDatum,
    GroundStateTrainingConfig,
    create_train_state_from_molecule,
    ground_state_mse_loss,
    ground_state_mse_loss_pointwise_dataset,
    make_ground_state_loss_and_grad,
    make_ground_state_train_step,
    predict_ground_state_total_energy,
    load_params_checkpoint,
    save_params_checkpoint,
)
from td_graddft.training.targets import (
    _predict_ground_state_total_energy_from_molecule,
    _resolve_training_molecule_and_info_with_mode,
)

HARTREE_TO_EV = 27.211386245988
_DEFAULT_SEMILOCAL_XC = ("lda_x", "gga_x_b88", "lda_c_vwn_rpa", "gga_c_lyp")
_TRAIN_SCF_SAFETY_MAX_CYCLE = 512
_PREDICTION_SCF_STABILITY_FIELDS = (
    "scf_stable",
    "scf_energy_spread_ev",
    "scf_min_energy_h",
    "scf_max_energy_h",
    "scf_converged_all",
    "scf_max_cycles",
    "scf_max_selected_rms_density",
)


@dataclass(frozen=True)
class ReferencePoint:
    r_angstrom: float
    molecule: Any
    reference_energy_h: float
    mean_field_energy_h: float
    correlation_energy_h: float
    perturbative_corr_h: float
    reference_method: str
    target_density_matrix: Any


class RunLogger:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def log(self, message: str) -> None:
        stamp = time.strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{stamp}] {message}"
        print(line, flush=True)
        with self.path.open("a", encoding="utf-8") as handle:
            handle.write(line + "\n")


def build_diatomic_atom(atom_symbol: str, r_angstrom: float) -> str:
    half = 0.5 * float(r_angstrom)
    symbol = str(atom_symbol)
    return f"{symbol} 0.0 0.0 {-half:.10f}; {symbol} 0.0 0.0 {half:.10f}"


def build_n2_atom(r_angstrom: float) -> str:
    return build_diatomic_atom("N", r_angstrom)


def _xlsx_first_sheet_rows(path: Path):
    import zipfile
    import xml.etree.ElementTree as ET

    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _text(node) -> str:
        return "".join(text.text or "" for text in node.findall(".//main:t", ns))

    with zipfile.ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [_text(item) for item in shared_root.findall("main:si", ns)]

        sheet_path = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in archive.namelist() and "xl/_rels/workbook.xml.rels" in archive.namelist():
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            first_sheet = workbook.find(".//main:sheets/main:sheet", ns)
            rel_id = first_sheet.get(f"{{{ns['rel']}}}id") if first_sheet is not None else None
            rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            for rel in rels.findall("pkgrel:Relationship", ns):
                if rel.get("Id") == rel_id:
                    target = str(rel.get("Target"))
                    sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
                    break

        sheet = ET.fromstring(archive.read(sheet_path))
        for row in sheet.findall(".//main:sheetData/main:row", ns):
            values: list[str | None] = []
            for cell in row.findall("main:c", ns):
                value_node = cell.find("main:v", ns)
                if value_node is None:
                    inline = cell.find("main:is", ns)
                    values.append(_text(inline) if inline is not None else None)
                    continue
                value = value_node.text
                if cell.get("t") == "s" and value is not None:
                    value = shared_strings[int(value)]
                values.append(value)
            yield tuple(values)


def load_graddft_dissociation_targets(
    path: str | Path,
    *,
    r_values: list[float] | np.ndarray,
) -> dict[float, float]:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"GradDFT dissociation file not found: {source}")
    if source.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            raw_rows = _xlsx_first_sheet_rows(source)
        else:
            workbook = load_workbook(source, data_only=True, read_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            raw_rows = sheet.iter_rows(values_only=True)
    else:
        handle = source.open(newline="", encoding="utf-8")
        reader = csv.reader(handle)
        raw_rows = reader

    rows: list[tuple[float, float]] = []
    try:
        for row in raw_rows:
            if row is None or len(row) < 2:
                continue
            try:
                r_value = float(row[0])
                energy = float(row[1])
            except (TypeError, ValueError):
                continue
            rows.append((r_value, energy))
    finally:
        close = getattr(locals().get("handle", None), "close", None)
        if callable(close):
            close()
    if len(rows) < 2:
        raise ValueError(f"Need at least two GradDFT dissociation points in {source}")

    rows = sorted(rows, key=lambda item: item[0])
    r_table = np.asarray([row[0] for row in rows], dtype=np.float64)
    e_table = np.asarray([row[1] for row in rows], dtype=np.float64)
    query = np.asarray([float(value) for value in r_values], dtype=np.float64)
    if np.any(query < r_table[0]) or np.any(query > r_table[-1]):
        raise ValueError(
            "Requested R grid is outside GradDFT data range: "
            f"requested [{float(query.min()):.8g}, {float(query.max()):.8g}], "
            f"available [{float(r_table[0]):.8g}, {float(r_table[-1]):.8g}]"
        )
    targets: dict[float, float] = {}
    for r_value in query:
        matches = np.nonzero(np.isclose(r_table, float(r_value), rtol=0.0, atol=1e-8))[0]
        if len(matches) == 0:
            raise ValueError(
                "No exact GradDFT target for "
                f"R={float(r_value):.8g} in {source}; choose one of the tabulated R values."
            )
        targets[float(r_value)] = float(e_table[int(matches[0])])
    return targets


def _move_scf_to_gpu(mf: Any) -> Any:
    try:
        import gpu4pyscf  # noqa: F401
    except ModuleNotFoundError as exc:
        raise RuntimeError("--reference-scf-device gpu requires gpu4pyscf.") from exc
    to_gpu = getattr(mf, "to_gpu", None)
    if not callable(to_gpu):
        raise RuntimeError("gpu4pyscf did not expose a to_gpu() method on this SCF object.")
    return to_gpu()


def _move_scf_to_cpu(mf: Any) -> Any:
    to_cpu = getattr(mf, "to_cpu", None)
    return to_cpu() if callable(to_cpu) else mf


def _run_rhf(mol: gto.Mole, *, max_cycle: int, conv_tol: float) -> scf.hf.RHF:
    attempts = (
        dict(init_guess="minao", damping=0.20, level_shift=0.0, newton=False),
        dict(init_guess="atom", damping=0.35, level_shift=0.5, newton=False),
        dict(init_guess="atom", damping=0.0, level_shift=0.0, newton=True),
    )
    last = None
    for opts in attempts:
        mf = scf.RHF(mol)
        mf.conv_tol = float(conv_tol)
        mf.max_cycle = int(max_cycle)
        mf.init_guess = str(opts["init_guess"])
        mf.damping = float(opts["damping"])
        mf.level_shift = float(opts["level_shift"])
        mf.diis_start_cycle = 1
        if bool(opts["newton"]):
            mf = mf.newton()
            mf.conv_tol = float(conv_tol)
            mf.max_cycle = int(max_cycle)
        mf.kernel()
        last = mf
        if bool(mf.converged):
            return mf
    raise RuntimeError(f"RHF did not converge; last E={getattr(last, 'e_tot', None)}")


def _run_rks(
    mol: gto.Mole,
    *,
    xc: str,
    grids_level: int,
    max_cycle: int,
    conv_tol: float,
    reference_scf_device: str,
    jk_backend: str,
) -> dft.rks.RKS:
    attempts = (
        dict(init_guess="minao", damping=0.20, level_shift=0.0, newton=False),
        dict(init_guess="atom", damping=0.35, level_shift=0.5, newton=False),
        dict(init_guess="atom", damping=0.0, level_shift=0.0, newton=True),
    )
    last = None
    for opts in attempts:
        mf = dft.RKS(mol)
        mf.xc = str(xc)
        if str(jk_backend) == "df":
            mf = mf.density_fit()
        mf.grids.level = int(grids_level)
        mf.conv_tol = float(conv_tol)
        mf.max_cycle = int(max_cycle)
        mf.init_guess = str(opts["init_guess"])
        mf.damping = float(opts["damping"])
        mf.level_shift = float(opts["level_shift"])
        mf.diis_start_cycle = 1
        if bool(opts["newton"]):
            mf = mf.newton()
            mf.xc = str(xc)
            mf.conv_tol = float(conv_tol)
            mf.max_cycle = int(max_cycle)
        if str(reference_scf_device) == "gpu":
            mf = _move_scf_to_gpu(mf)
        mf.kernel()
        last = mf
        if bool(mf.converged):
            return _move_scf_to_cpu(mf)
    raise RuntimeError(f"RKS did not converge; last E={getattr(last, 'e_tot', None)}")


def compute_ccsdt_energy(
    mol: gto.Mole,
    *,
    rhf_max_cycle: int,
    rhf_conv_tol: float,
    ccsd_max_cycle: int,
    ccsd_conv_tol: float,
) -> tuple[float, float, float, float, np.ndarray]:
    mf = _run_rhf(mol, max_cycle=int(rhf_max_cycle), conv_tol=float(rhf_conv_tol))
    solver = cc.CCSD(mf)
    solver.max_cycle = int(ccsd_max_cycle)
    solver.conv_tol = float(ccsd_conv_tol)
    eccsd, _, _ = solver.kernel()
    if not bool(solver.converged):
        raise RuntimeError("CCSD did not converge")
    et = solver.ccsd_t()
    ccsd_dm_ao = np.asarray(solver.make_rdm1(ao_repr=True), dtype=np.float64)
    total = float(mf.e_tot + eccsd + et)
    return total, float(mf.e_tot), float(eccsd), float(et), ccsd_dm_ao


def compute_casscf_energy(
    mol: gto.Mole,
    *,
    active_space: str,
    active_labels: list[str],
    ncas: int,
    nelecas: int,
    avas_threshold: float,
    rhf_max_cycle: int,
    rhf_conv_tol: float,
    casscf_max_cycle: int,
    casscf_conv_tol: float,
    include_nevpt2: bool,
) -> tuple[float, float, float, float, np.ndarray]:
    from pyscf import mcscf

    mf = _run_rhf(mol, max_cycle=int(rhf_max_cycle), conv_tol=float(rhf_conv_tol))
    mo_coeff = None
    if str(active_space) == "avas":
        from pyscf.mcscf import avas

        avas_ncas, avas_nelecas, mo_coeff = avas.avas(
            mf,
            list(active_labels),
            threshold=float(avas_threshold),
            canonicalize=True,
        )
        if int(avas_ncas) < int(ncas) or int(avas_nelecas) < int(nelecas):
            raise RuntimeError(
                "AVAS active space mismatch: "
                f"expected at least ({int(nelecas)}e,{int(ncas)}o), "
                f"got ({int(avas_nelecas)}e,{int(avas_ncas)}o)"
            )
    mc = mcscf.CASSCF(mf, int(ncas), int(nelecas))
    mc.conv_tol = float(casscf_conv_tol)
    if hasattr(mc, "max_cycle_macro"):
        mc.max_cycle_macro = int(casscf_max_cycle)
    else:
        mc.max_cycle = int(casscf_max_cycle)
    mc.kernel(mo_coeff)
    if not bool(mc.converged):
        raise RuntimeError(f"CASSCF did not converge; last E={getattr(mc, 'e_tot', None)}")
    casscf_total = float(mc.e_tot)
    perturbative_corr = 0.0
    if bool(include_nevpt2):
        from pyscf import mrpt

        perturbative_corr = float(mrpt.NEVPT(mc).kernel())
    total = casscf_total + perturbative_corr
    dm_ao = np.asarray(mc.make_rdm1(), dtype=np.float64)
    return total, float(mf.e_tot), float(casscf_total - mf.e_tot), perturbative_corr, dm_ao


def build_reference_point(
    r_angstrom: float,
    *,
    args: argparse.Namespace,
    graddft_target_energy_h: float | None = None,
) -> ReferencePoint:
    mol = gto.M(
        atom=build_diatomic_atom(str(args.atom_symbol), float(r_angstrom)),
        basis=str(args.basis),
        unit="Angstrom",
        spin=0,
        charge=0,
        cart=True,
        verbose=0,
    )
    if str(args.reference_method) == "graddft_data":
        if graddft_target_energy_h is None:
            raise ValueError("graddft_target_energy_h is required for graddft_data references")
        reference_energy = float(graddft_target_energy_h)
        corr_energy = 0.0
        perturbative_corr = 0.0
        reference_dm_ao = None
    elif str(args.reference_method) == "ccsd_t":
        reference_energy, mf_energy, corr_energy, perturbative_corr, reference_dm_ao = (
            compute_ccsdt_energy(
                mol,
                rhf_max_cycle=int(args.rhf_max_cycle),
                rhf_conv_tol=float(args.rhf_conv_tol),
                ccsd_max_cycle=int(args.ccsd_max_cycle),
                ccsd_conv_tol=float(args.ccsd_conv_tol),
            )
        )
    else:
        reference_energy, mf_energy, corr_energy, perturbative_corr, reference_dm_ao = (
            compute_casscf_energy(
                mol,
                active_space=str(args.active_space),
                active_labels=list(args.active_labels),
                ncas=int(args.ncas),
                nelecas=int(args.nelecas),
                avas_threshold=float(args.avas_threshold),
                rhf_max_cycle=int(args.rhf_max_cycle),
                rhf_conv_tol=float(args.rhf_conv_tol),
                casscf_max_cycle=int(args.casscf_max_cycle),
                casscf_conv_tol=float(args.casscf_conv_tol),
                include_nevpt2=(str(args.reference_method) == "casscf_nevpt2"),
            )
        )
    mf_ref = _run_rks(
        mol,
        xc=str(args.xc),
        grids_level=int(args.grids_level),
        max_cycle=int(args.reference_scf_max_cycle),
        conv_tol=float(args.reference_scf_conv_tol),
        reference_scf_device=str(args.reference_scf_device),
        jk_backend=str(args.jk_backend),
    )
    reference = restricted_reference_from_pyscf(
        mf_ref,
        compute_local_hfx_features=(
            str(args.input_feature_mode) == "canonical" or bool(args.include_hfx_channel)
        ),
        compute_local_hfx_aux=(
            str(args.input_feature_mode) == "canonical" or bool(args.include_hfx_channel)
        ),
        compute_local_pt2_features=bool(args.include_pt2_channel),
        hfx_nu_storage=str(args.hfx_nu_storage),
        jk_backend=str(args.jk_backend),
    )
    if str(args.reference_method) == "graddft_data":
        mf_energy = float(mf_ref.e_tot)
        corr_energy = float(reference_energy - mf_energy)
        reference_dm_ao = np.asarray(mf_ref.make_rdm1(), dtype=np.float64)
    return ReferencePoint(
        r_angstrom=float(r_angstrom),
        molecule=reference,
        reference_energy_h=float(reference_energy),
        mean_field_energy_h=float(mf_energy),
        correlation_energy_h=float(corr_energy),
        perturbative_corr_h=float(perturbative_corr),
        reference_method=str(args.reference_method),
        target_density_matrix=jnp.asarray(reference_dm_ao, dtype=jnp.float64),
    )


def build_training_data(
    points: list[ReferencePoint],
    *,
    density_constraint_weight: float,
    density_matrix_constraint_weight: float,
) -> tuple[GroundStateDatum, ...]:
    return tuple(
        GroundStateDatum.from_parts(
            point.molecule,
            core=GroundStateCoreDatum(
                target_total_energy=jnp.asarray(point.reference_energy_h, dtype=jnp.float64),
                target_density_matrix=jnp.asarray(point.target_density_matrix),
                density_constraint_weight=float(density_constraint_weight),
                density_matrix_constraint_weight=float(density_matrix_constraint_weight),
            ),
        )
        for point in points
    )


def _cache_base_group_name(args: argparse.Namespace) -> str:
    pt2 = "pt2" if bool(args.include_pt2_channel) else "nopt2"
    hfx = "hfx" if bool(args.include_hfx_channel) else "nohfx"
    active = "none"
    if str(args.reference_method) in {"casscf", "casscf_nevpt2"}:
        labels = "-".join(str(label).replace(" ", "") for label in args.active_labels)
        active = (
            f"{str(args.active_space)}/ncas={int(args.ncas)}/"
            f"nelecas={int(args.nelecas)}/labels={labels}/"
            f"thr={float(args.avas_threshold):.3g}"
        )
    return (
        f"{str(args.atom_symbol).lower()}2_ground/ref={str(args.reference_method)}/"
        f"active={active}/"
        f"basis={str(args.basis).replace('/', '_')}/"
        f"grid={int(args.grids_level)}/"
        f"jk={str(args.jk_backend)}/"
        f"{pt2}/"
        f"{hfx}/"
        f"hfmode={str(args.ground_state_hf_mode)}/"
        f"pt2mode={str(args.pt2_channel_mode) if bool(args.include_pt2_channel) else 'none'}_"
        f"{str(args.ground_state_pt2_mode)}/"
        f"hfxnu={str(args.hfx_nu_storage)}"
    )


def _cache_group_name(args: argparse.Namespace) -> str:
    r_values = getattr(args, "r_values", None)
    if r_values is not None:
        values = "-".join(f"{float(value):.8g}" for value in r_values)
        return f"{_cache_base_group_name(args)}/rvalues={values}"
    return (
        f"{_cache_base_group_name(args)}/"
        f"r={float(args.r_min):.8g}-{float(args.r_max):.8g}/"
        f"n={int(args.train_points)}"
    )


def _point_cache_group_name(args: argparse.Namespace, r_value: float) -> str:
    return f"{_cache_base_group_name(args)}/points/r={float(r_value):.8g}"


def _reference_label(args: argparse.Namespace) -> str:
    method = str(args.reference_method)
    if method == "ccsd_t":
        return "CCSD(T)"
    if method == "graddft_data":
        return "GradDFT MR-ccCA"
    label = f"CASSCF({int(args.nelecas)}e,{int(args.ncas)}o)"
    return f"{label}+NEVPT2" if method == "casscf_nevpt2" else label


def _write_reference_point(group: Any, point: ReferencePoint) -> None:
    group.attrs["r_angstrom"] = float(point.r_angstrom)
    group.attrs["reference_energy_h"] = float(point.reference_energy_h)
    group.attrs["mean_field_energy_h"] = float(point.mean_field_energy_h)
    group.attrs["correlation_energy_h"] = float(point.correlation_energy_h)
    group.attrs["perturbative_corr_h"] = float(point.perturbative_corr_h)
    group.attrs["reference_method"] = str(point.reference_method)
    group.attrs["ccsdt_energy_h"] = float(point.reference_energy_h)
    group.attrs["rhf_energy_h"] = float(point.mean_field_energy_h)
    group.attrs["ccsd_corr_h"] = float(point.correlation_energy_h)
    group.attrs["triples_corr_h"] = float(point.perturbative_corr_h)
    group.create_dataset(
        "target_density_matrix",
        data=np.asarray(point.target_density_matrix, dtype=np.float64),
    )
    write_restricted_molecule(group.require_group("molecule"), point.molecule)


def _read_reference_point(
    group: Any,
    *,
    hfx_nu_storage: str = "array",
) -> ReferencePoint:
    molecule = read_restricted_molecule(
        group["molecule"],
        hfx_nu_storage=("chunked" if str(hfx_nu_storage) == "chunked" else "array"),
    )
    target_density_matrix = (
        group["target_density_matrix"][()]
        if "target_density_matrix" in group
        else np.asarray(molecule.rdm1).sum(axis=0)
    )
    return ReferencePoint(
        r_angstrom=float(group.attrs["r_angstrom"]),
        molecule=molecule,
        reference_energy_h=float(
            group.attrs["reference_energy_h"]
            if "reference_energy_h" in group.attrs
            else group.attrs["ccsdt_energy_h"]
        ),
        mean_field_energy_h=float(
            group.attrs["mean_field_energy_h"]
            if "mean_field_energy_h" in group.attrs
            else group.attrs["rhf_energy_h"]
        ),
        correlation_energy_h=float(
            group.attrs["correlation_energy_h"]
            if "correlation_energy_h" in group.attrs
            else group.attrs["ccsd_corr_h"]
        ),
        perturbative_corr_h=float(
            group.attrs["perturbative_corr_h"]
            if "perturbative_corr_h" in group.attrs
            else group.attrs["triples_corr_h"]
        ),
        reference_method=str(group.attrs.get("reference_method", "ccsd_t")),
        target_density_matrix=jnp.asarray(target_density_matrix),
    )


def _has_hdf5_group(path: Path, group_name: str) -> bool:
    if not path.exists():
        return False
    with h5py.File(path, "r") as handle:
        return group_name in handle


def _save_reference_points_hdf5(path: Path, group_name: str, points: list[ReferencePoint]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with h5py.File(path, "a") as handle:
        if group_name in handle:
            del handle[group_name]
        group = handle.create_group(group_name)
        group.attrs["count"] = int(len(points))
        for idx, point in enumerate(points):
            _write_reference_point(group.create_group(f"point_{idx:04d}"), point)


def _load_reference_points_hdf5(
    path: Path,
    group_name: str,
    *,
    hfx_nu_storage: str = "array",
) -> list[ReferencePoint]:
    with h5py.File(path, "r") as handle:
        group = handle[group_name]
        count = int(group.attrs["count"])
        return [
            _read_reference_point(group[f"point_{idx:04d}"], hfx_nu_storage=hfx_nu_storage)
            for idx in range(count)
        ]


def _save_reference_point_hdf5(path: Path, group_name: str, point: ReferencePoint) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with h5py.File(path, "a") as handle:
        if group_name in handle:
            del handle[group_name]
        _write_reference_point(handle.create_group(group_name), point)


def _load_reference_point_hdf5(
    path: Path,
    group_name: str,
    *,
    hfx_nu_storage: str = "array",
) -> ReferencePoint:
    with h5py.File(path, "r") as handle:
        return _read_reference_point(handle[group_name], hfx_nu_storage=hfx_nu_storage)


def load_or_build_reference_point(
    r_value: float,
    *,
    args: argparse.Namespace,
    cache_path: Path | None,
    graddft_targets: dict[float, float],
) -> tuple[ReferencePoint, str]:
    point_group = _point_cache_group_name(args, float(r_value))
    if (
        cache_path is not None
        and _has_hdf5_group(cache_path, point_group)
        and not bool(args.rebuild_reference_cache)
    ):
        return _load_reference_point_hdf5(
            cache_path,
            point_group,
            hfx_nu_storage=str(args.hfx_nu_storage),
        ), "cache"
    point = build_reference_point(
        float(r_value),
        args=args,
        graddft_target_energy_h=graddft_targets.get(float(r_value)),
    )
    if cache_path is not None:
        _save_reference_point_hdf5(cache_path, point_group, point)
        if str(args.hfx_nu_storage) == "chunked":
            point = _load_reference_point_hdf5(
                cache_path,
                point_group,
                hfx_nu_storage="chunked",
            )
    return point, "built"


def _tree_add(left: Any | None, right: Any) -> Any:
    if left is None:
        return right
    return jax.tree_util.tree_map(lambda a, b: a + b, left, right)


def _tree_scale(tree: Any, scale: float) -> Any:
    return jax.tree_util.tree_map(lambda value: value * scale, tree)


def _metric_scalar(metrics: dict[str, Any], key: str, default: float = float("nan")) -> float:
    if key not in metrics:
        return default
    arr = jnp.asarray(metrics[key])
    if int(arr.size) <= 0:
        return default
    return float(jnp.mean(arr))


def _r_column_label(r_angstrom: float) -> str:
    text = f"{float(r_angstrom):.6g}".replace("-", "m").replace(".", "p")
    return f"r_{text}"


_STREAM_POINT_METRICS = (
    "loss",
    "energy_mae_h",
    "energy_signed_err_h",
    "predicted_energy_h",
    "density_mse",
    "density_matrix_mse",
    "grad_norm",
    "grad_abs_max",
    "nonfinite_grad_fraction",
    "scf_converged",
    "scf_cycles",
    "scf_selected_rms_density",
    "scf_best_rms_density",
)


def _add_stream_point_metrics(
    row: dict[str, float],
    *,
    labels: tuple[str, ...],
    losses: list[float],
    metric_rows: list[dict[str, float]],
) -> None:
    for label, loss_value, metrics in zip(labels, losses, metric_rows):
        row[f"loss_{label}"] = float(loss_value)
        for key in _STREAM_POINT_METRICS:
            if key == "loss":
                continue
            row[f"{key}_{label}"] = float(metrics.get(key, float("nan")))


def _tree_all_finite(tree: Any) -> bool:
    leaves = jax.tree_util.tree_leaves(tree)
    return all(bool(jnp.all(jnp.isfinite(jnp.asarray(leaf)))) for leaf in leaves)


def _tree_l2_norm(tree: Any) -> float:
    leaves = [jnp.asarray(leaf) for leaf in jax.tree_util.tree_leaves(tree)]
    if not leaves:
        return 0.0
    total = sum(jnp.sum(jnp.square(leaf)) for leaf in leaves)
    return float(jnp.sqrt(total))


def build_functional_and_training_config(
    args: argparse.Namespace,
) -> tuple[Any, GroundStateTrainingConfig]:
    functional = neural_xc.Functional(
        semilocal_xc=tuple(str(name) for name in args.semilocal_xc),
        hidden_dims=tuple(int(value) for value in args.hidden_dims),
        architecture=str(args.network_architecture),
        input_feature_mode=str(args.input_feature_mode),
        include_pt2_channel=bool(args.include_pt2_channel),
        include_hfx_channel=bool(args.include_hfx_channel),
        ground_state_hf_mode=str(args.ground_state_hf_mode),
        ground_state_pt2_mode=str(args.ground_state_pt2_mode),
        pt2_channel_mode=str(args.pt2_channel_mode),
        response_pt2_mode="approx",
        allow_experimental_jax_xc=bool(args.allow_experimental_jax_xc),
        name="neural_xc_n2_ccsdt_ground",
    )
    coefficient_prior = neural_xc.resolve_coefficient_prior_values(
        tuple(str(name) for name in args.semilocal_xc)
    )
    if coefficient_prior is not None and (
        bool(args.include_pt2_channel) or bool(args.include_hfx_channel)
    ):
        n_semilocal = len(tuple(str(name) for name in args.semilocal_xc))
        if len(coefficient_prior) == n_semilocal + 1:
            coefficient_prior = (
                tuple(coefficient_prior[:n_semilocal])
                + ((0.0,) if bool(args.include_pt2_channel) else ())
                + (tuple(coefficient_prior[n_semilocal:]) if bool(args.include_hfx_channel) else ())
            )
    training_config = GroundStateTrainingConfig.from_parts(
        core=GroundStateCoreTrainingConfig(
            mode=str(args.training_mode),
            energy_mse_weight=float(args.energy_mse_weight),
            energy_mae_weight=float(args.energy_mae_weight),
            energy_normalization=str(args.energy_normalization),
            coefficient_prior_weight=float(args.coefficient_prior_weight),
            coefficient_prior_values=coefficient_prior,
            scf_max_cycle=(
                _TRAIN_SCF_SAFETY_MAX_CYCLE
                if int(args.train_scf_max_cycle) <= 0
                else int(args.train_scf_max_cycle)
            ),
            scf_damping=float(args.train_scf_damping),
            scf_conv_tol_energy=args.train_scf_conv_tol_energy,
            scf_convergence_metric=str(args.train_scf_convergence_metric),
            scf_conv_tol_density=float(args.train_scf_conv_tol_density),
            scf_vxc_clip=float(args.train_scf_vxc_clip),
            scf_iterate_selection=str(args.scf_iterate_selection),
            scf_require_convergence=bool(args.scf_require_convergence),
            scf_gradient_mode=str(args.scf_gradient_mode),
            scf_implicit_diff_tolerance=float(args.scf_implicit_diff_tolerance),
            scf_implicit_diff_regularization=float(args.scf_implicit_diff_regularization),
        ),
    )
    return functional, training_config


def train_functional(
    points: list[ReferencePoint],
    *,
    args: argparse.Namespace,
    logger: RunLogger,
):
    train_data = build_training_data(
        points,
        density_constraint_weight=float(args.density_constraint_weight),
        density_matrix_constraint_weight=float(args.density_matrix_constraint_weight),
    )
    point_labels = tuple(_r_column_label(point.r_angstrom) for point in points)
    target_energies_h = tuple(float(point.reference_energy_h) for point in points)
    functional, training_config = build_functional_and_training_config(args)
    if int(args.lr_decay_every) > 0:
        lr_schedule = optax.exponential_decay(
            init_value=float(args.learning_rate),
            transition_steps=int(args.lr_decay_every),
            decay_rate=float(args.lr_decay_factor),
            staircase=True,
        )
        base_optimizer = optax.adam(lr_schedule)
    else:
        lr_schedule = None
        base_optimizer = optax.adam(float(args.learning_rate))
    state = create_train_state_from_molecule(
        functional,
        jax.random.PRNGKey(int(args.seed)),
        points[0].molecule,
        base_optimizer,
    )
    if bool(args.stream_train):
        loss_and_grad = make_ground_state_loss_and_grad(
            functional,
            training_config=training_config,
        )
        eval_single = lambda params, datum: ground_state_mse_loss(  # noqa: E731
            params,
            functional,
            datum,
            training_config=training_config,
        )
        if bool(args.jit_train):
            loss_and_grad = jax.jit(loss_and_grad)
        if bool(args.jit_eval):
            eval_single = jax.jit(eval_single)

        def _eval_all(params):
            losses = []
            maes = []
            density_mses = []
            density_penalties = []
            density_matrix_mses = []
            density_matrix_penalties = []
            point_metric_rows = []
            for idx, datum in enumerate(train_data):
                loss_val, metrics_val = eval_single(params, datum)
                loss_scalar = float(loss_val)
                energy_mae = _metric_scalar(metrics_val, "energy_mae")
                density_mse = _metric_scalar(metrics_val, "density_mse")
                density_penalty = _metric_scalar(metrics_val, "density_penalty")
                density_matrix_mse = _metric_scalar(metrics_val, "density_matrix_mse")
                density_matrix_penalty = _metric_scalar(metrics_val, "density_matrix_penalty")
                predicted_energy = _metric_scalar(metrics_val, "predicted_total_energies")
                losses.append(loss_scalar)
                maes.append(energy_mae)
                density_mses.append(density_mse)
                density_penalties.append(density_penalty)
                density_matrix_mses.append(density_matrix_mse)
                density_matrix_penalties.append(density_matrix_penalty)
                point_metric_rows.append(
                    {
                        "energy_mae_h": energy_mae,
                        "energy_signed_err_h": predicted_energy - target_energies_h[idx],
                        "predicted_energy_h": predicted_energy,
                        "density_mse": density_mse,
                        "density_matrix_mse": density_matrix_mse,
                        "scf_converged": _metric_scalar(metrics_val, "scf_converged"),
                        "scf_cycles": _metric_scalar(metrics_val, "scf_cycles"),
                        "scf_selected_rms_density": _metric_scalar(
                            metrics_val,
                            "scf_selected_rms_density",
                        ),
                        "scf_best_rms_density": _metric_scalar(
                            metrics_val,
                            "scf_best_rms_density",
                        ),
                    }
                )
            row = {
                "loss": float(np.mean(losses)),
                "energy_mae_h": float(np.mean(maes)),
                "density_mse": float(np.mean(density_mses)),
                "density_penalty": float(np.mean(density_penalties)),
                "density_matrix_mse": float(np.mean(density_matrix_mses)),
                "density_matrix_penalty": float(np.mean(density_matrix_penalties)),
            }
            _add_stream_point_metrics(
                row,
                labels=point_labels,
                losses=losses,
                metric_rows=point_metric_rows,
            )
            return row

        initial_eval = _eval_all(state.params)
        best_params = state.params
        min_loss = initial_eval["loss"]
        min_loss_step = 0
        rows = [
            {
                "step": 0,
                **initial_eval,
                "grad_norm": float("nan"),
                "grad_abs_max": float("nan"),
                "param_update_norm": float("nan"),
                "nonfinite_grad_fraction": 0.0,
                "lr": float(args.learning_rate),
            }
        ]
        logger.log(
            "[train] "
            f"steps={int(args.steps)} lr={float(args.learning_rate):.6g} "
            f"lr_decay_every={int(args.lr_decay_every)} lr_decay_factor={float(args.lr_decay_factor):.6g} "
            f"include_pt2_channel={bool(args.include_pt2_channel)} "
            f"include_hfx_channel={bool(args.include_hfx_channel)} "
            "train_step_mode=stream_single_geometry"
        )
        t0 = time.perf_counter()
        for step in range(1, int(args.steps) + 1):
            prev_state = state
            grad_sum = None
            losses = []
            maes = []
            density_mses = []
            density_penalties = []
            density_matrix_mses = []
            density_matrix_penalties = []
            grad_norms = []
            grad_abs_maxes = []
            nonfinite_fracs = []
            point_metric_rows = []
            for idx, datum in enumerate(train_data):
                loss_val, metrics, grads = loss_and_grad(state.params, datum)
                loss_scalar = float(loss_val)
                energy_mae = _metric_scalar(metrics, "energy_mae")
                density_mse = _metric_scalar(metrics, "density_mse")
                density_matrix_mse = _metric_scalar(metrics, "density_matrix_mse")
                grad_norm = _metric_scalar(metrics, "grad_norm")
                grad_abs_max = _metric_scalar(metrics, "grad_abs_max")
                nonfinite_frac = _metric_scalar(metrics, "nonfinite_grad_fraction", 0.0)
                predicted_energy = _metric_scalar(metrics, "predicted_total_energies")
                losses.append(loss_scalar)
                maes.append(energy_mae)
                density_mses.append(density_mse)
                density_penalties.append(_metric_scalar(metrics, "density_penalty"))
                density_matrix_mses.append(density_matrix_mse)
                density_matrix_penalties.append(_metric_scalar(metrics, "density_matrix_penalty"))
                grad_norms.append(grad_norm)
                grad_abs_maxes.append(grad_abs_max)
                nonfinite_fracs.append(nonfinite_frac)
                point_metric_rows.append(
                    {
                        "energy_mae_h": energy_mae,
                        "energy_signed_err_h": predicted_energy - target_energies_h[idx],
                        "predicted_energy_h": predicted_energy,
                        "density_mse": density_mse,
                        "density_matrix_mse": density_matrix_mse,
                        "grad_norm": grad_norm,
                        "grad_abs_max": grad_abs_max,
                        "nonfinite_grad_fraction": nonfinite_frac,
                        "scf_converged": _metric_scalar(metrics, "scf_converged"),
                        "scf_cycles": _metric_scalar(metrics, "scf_cycles"),
                        "scf_selected_rms_density": _metric_scalar(
                            metrics,
                            "scf_selected_rms_density",
                        ),
                        "scf_best_rms_density": _metric_scalar(
                            metrics,
                            "scf_best_rms_density",
                        ),
                    }
                )
                grad_sum = _tree_add(grad_sum, grads)
            grads_avg = _tree_scale(grad_sum, 1.0 / max(1, len(train_data)))
            state = state.apply_gradients(grads=grads_avg)
            param_delta = jax.tree_util.tree_map(
                lambda new, old: new - old,
                state.params,
                prev_state.params,
            )
            reverted = False
            if not _tree_all_finite(state.params):
                state = prev_state
                reverted = True
                logger.log(f"[train] non-finite params detected at step {step}; reverted update")
            loss = float(np.mean(losses))
            if loss < min_loss:
                min_loss = loss
                min_loss_step = step
                best_params = prev_state.params
            lr = float(lr_schedule(step - 1)) if lr_schedule is not None else float(args.learning_rate)
            row = {
                "step": step,
                "loss": loss,
                "energy_mae_h": float(np.mean(maes)),
                "density_mse": float(np.mean(density_mses)),
                "density_penalty": float(np.mean(density_penalties)),
                "density_matrix_mse": float(np.mean(density_matrix_mses)),
                "density_matrix_penalty": float(np.mean(density_matrix_penalties)),
                "grad_norm": float(np.mean(grad_norms)),
                "grad_abs_max": float(np.max(grad_abs_maxes)),
                "param_update_norm": 0.0 if reverted else float(_tree_l2_norm(param_delta)),
                "nonfinite_grad_fraction": float(np.mean(nonfinite_fracs)),
                "lr": lr,
            }
            _add_stream_point_metrics(
                row,
                labels=point_labels,
                losses=losses,
                metric_rows=point_metric_rows,
            )
            rows.append(row)
            if step == 1 or step % int(args.log_every) == 0 or step == int(args.steps):
                logger.log(
                    "[train] "
                    f"step={step:4d}/{int(args.steps):4d} "
                    f"loss={row['loss']:.8e} "
                    f"energy_mae={row['energy_mae_h']:.8e} "
                    f"density_mse={row['density_mse']:.8e} "
                    f"dm_mse={row['density_matrix_mse']:.8e} "
                    f"grad_norm={row['grad_norm']:.8e} "
                    f"update_norm={row['param_update_norm']:.8e} "
                    f"lr={row['lr']:.8e}"
                )
        elapsed = time.perf_counter() - t0
        final_eval = _eval_all(state.params)
        if final_eval["loss"] < min_loss:
            min_loss = final_eval["loss"]
            min_loss_step = int(args.steps)
            best_params = state.params
        logger.log(
            "[train] done "
            f"final_loss={final_eval['loss']:.8e} min_loss={min_loss:.8e}@{min_loss_step} "
            f"elapsed_s={elapsed:.2f} param_norm={_tree_l2_norm(state.params):.8e}"
        )
        return {
            "functional": functional,
            "training_config": training_config,
            "params": state.params,
            "best_params": best_params,
            "history": rows,
            "elapsed_s": elapsed,
            "final_loss": final_eval["loss"],
            "final_energy_mae_h": final_eval["energy_mae_h"],
            "min_loss": min_loss,
            "min_loss_step": min_loss_step,
        }

    train_step = make_ground_state_train_step(
        functional,
        training_config=training_config,
        loss_fn=ground_state_mse_loss_pointwise_dataset,
    )
    eval_fn = lambda params: ground_state_mse_loss_pointwise_dataset(  # noqa: E731
        params,
        functional,
        train_data,
        training_config=training_config,
    )
    step_fn = lambda current_state: train_step(current_state, train_data)  # noqa: E731
    compiled_eval = jax.jit(eval_fn) if bool(args.jit_eval) else eval_fn
    compiled_step = step_fn
    train_step_mode = "eager"
    if bool(args.jit_train):
        candidate = jax.jit(step_fn)
        try:
            _ = candidate.lower(state).compile()
            compiled_step = candidate
            train_step_mode = "jit"
        except Exception as exc:
            logger.log(f"[train] jit compilation failed: {exc!r}")

    initial_loss, initial_metrics = compiled_eval(state.params)
    best_params = state.params
    min_loss = float(initial_loss)
    min_loss_step = 0
    rows = [
        {
            "step": 0,
            "loss": float(initial_loss),
            "energy_mae_h": _metric_scalar(initial_metrics, "energy_mae"),
            "density_mse": _metric_scalar(initial_metrics, "density_mse"),
            "density_penalty": _metric_scalar(initial_metrics, "density_penalty"),
            "density_matrix_mse": _metric_scalar(initial_metrics, "density_matrix_mse"),
            "density_matrix_penalty": _metric_scalar(initial_metrics, "density_matrix_penalty"),
            "grad_norm": float("nan"),
            "grad_abs_max": float("nan"),
            "param_update_norm": float("nan"),
            "nonfinite_grad_fraction": 0.0,
            "lr": float(args.learning_rate),
        }
    ]
    logger.log(
        "[train] "
        f"steps={int(args.steps)} lr={float(args.learning_rate):.6g} "
        f"lr_decay_every={int(args.lr_decay_every)} lr_decay_factor={float(args.lr_decay_factor):.6g} "
        f"include_pt2_channel={bool(args.include_pt2_channel)} "
        f"include_hfx_channel={bool(args.include_hfx_channel)} "
        f"train_step_mode={train_step_mode}"
    )

    t0 = time.perf_counter()
    for step in range(1, int(args.steps) + 1):
        prev_state = state
        state, metrics = compiled_step(state)
        reverted = False
        if not _tree_all_finite(state.params):
            state = prev_state
            reverted = True
            logger.log(f"[train] non-finite params detected at step {step}; reverted update")
        loss = _metric_scalar(metrics, "loss")
        if step >= 2 and loss < min_loss:
            min_loss = loss
            min_loss_step = step - 1
            best_params = prev_state.params
        lr = float(lr_schedule(step - 1)) if lr_schedule is not None else float(args.learning_rate)
        row = {
            "step": step,
            "loss": loss,
            "energy_mae_h": _metric_scalar(metrics, "energy_mae"),
            "density_mse": _metric_scalar(metrics, "density_mse"),
            "density_penalty": _metric_scalar(metrics, "density_penalty"),
            "density_matrix_mse": _metric_scalar(metrics, "density_matrix_mse"),
            "density_matrix_penalty": _metric_scalar(metrics, "density_matrix_penalty"),
            "grad_norm": _metric_scalar(metrics, "grad_norm"),
            "grad_abs_max": _metric_scalar(metrics, "grad_abs_max"),
            "param_update_norm": 0.0 if reverted else _metric_scalar(metrics, "param_update_norm"),
            "nonfinite_grad_fraction": _metric_scalar(metrics, "nonfinite_grad_fraction", 0.0),
            "scf_converged_fraction": _metric_scalar(metrics, "scf_converged_fraction"),
            "scf_cycles_mean": _metric_scalar(metrics, "scf_cycles_mean"),
            "scf_cycles_max": _metric_scalar(metrics, "scf_cycles_max"),
            "scf_selected_rms_max": _metric_scalar(metrics, "scf_selected_rms_max"),
            "lr": lr,
        }
        rows.append(row)
        if step == 1 or step % int(args.log_every) == 0 or step == int(args.steps):
            logger.log(
                "[train] "
                f"step={step:4d}/{int(args.steps):4d} "
                f"loss={row['loss']:.8e} "
                f"energy_mae={row['energy_mae_h']:.8e} "
                f"density_mse={row['density_mse']:.8e} "
                f"dm_mse={row['density_matrix_mse']:.8e} "
                f"scf_conv_frac={row['scf_converged_fraction']:.6f} "
                f"scf_cycles_max={row['scf_cycles_max']:.6f} "
                f"scf_selected_rms_max={row['scf_selected_rms_max']:.8e} "
                f"grad_norm={row['grad_norm']:.8e} "
                f"update_norm={row['param_update_norm']:.8e} "
                f"lr={row['lr']:.8e}"
            )
    elapsed = time.perf_counter() - t0
    final_loss, final_metrics = compiled_eval(state.params)
    if float(final_loss) < min_loss:
        min_loss = float(final_loss)
        min_loss_step = int(args.steps)
        best_params = state.params
    logger.log(
        "[train] done "
        f"final_loss={float(final_loss):.8e} min_loss={min_loss:.8e}@{min_loss_step} "
        f"elapsed_s={elapsed:.2f} param_norm={_tree_l2_norm(state.params):.8e}"
    )
    return {
        "functional": functional,
        "training_config": training_config,
        "params": state.params,
        "best_params": best_params,
        "history": rows,
        "elapsed_s": elapsed,
        "final_loss": float(final_loss),
        "final_energy_mae_h": _metric_scalar(final_metrics, "energy_mae"),
        "min_loss": min_loss,
        "min_loss_step": min_loss_step,
    }


def write_history_csv(path: Path, rows: list[dict[str, float]]) -> None:
    fieldnames = sorted({key for row in rows for key in row})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_per_point_history_csv(
    path: Path,
    rows: list[dict[str, float]],
    points: list[ReferencePoint],
) -> None:
    labels = tuple(_r_column_label(point.r_angstrom) for point in points)
    fieldnames = [
        "step",
        "r_angstrom",
        "target_energy_h",
        "loss",
        "energy_mae_h",
        "energy_abs_err_ev",
        "energy_signed_err_h",
        "predicted_energy_h",
        "density_mse",
        "density_matrix_mse",
        "grad_norm",
        "grad_abs_max",
        "nonfinite_grad_fraction",
        "scf_converged",
        "scf_cycles",
        "scf_selected_rms_density",
        "scf_best_rms_density",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            for label, point in zip(labels, points):
                energy_mae_h = row.get(f"energy_mae_h_{label}", float("nan"))
                writer.writerow(
                    {
                        "step": int(row["step"]),
                        "r_angstrom": float(point.r_angstrom),
                        "target_energy_h": float(point.reference_energy_h),
                        "loss": row.get(f"loss_{label}", float("nan")),
                        "energy_mae_h": energy_mae_h,
                        "energy_abs_err_ev": float(energy_mae_h) * HARTREE_TO_EV,
                        "energy_signed_err_h": row.get(
                            f"energy_signed_err_h_{label}",
                            float("nan"),
                        ),
                        "predicted_energy_h": row.get(
                            f"predicted_energy_h_{label}",
                            float("nan"),
                        ),
                        "density_mse": row.get(f"density_mse_{label}", float("nan")),
                        "density_matrix_mse": row.get(
                            f"density_matrix_mse_{label}",
                            float("nan"),
                        ),
                        "grad_norm": row.get(f"grad_norm_{label}", float("nan")),
                        "grad_abs_max": row.get(f"grad_abs_max_{label}", float("nan")),
                        "nonfinite_grad_fraction": row.get(
                            f"nonfinite_grad_fraction_{label}",
                            float("nan"),
                        ),
                        "scf_converged": row.get(f"scf_converged_{label}", float("nan")),
                        "scf_cycles": row.get(f"scf_cycles_{label}", float("nan")),
                        "scf_selected_rms_density": row.get(
                            f"scf_selected_rms_density_{label}",
                            float("nan"),
                        ),
                        "scf_best_rms_density": row.get(
                            f"scf_best_rms_density_{label}",
                            float("nan"),
                        ),
                    }
                )


def prediction_rows_to_eval_history(
    pred_rows: list[dict[str, float]],
) -> list[dict[str, float]]:
    if not pred_rows:
        raise ValueError("At least one prediction row is required.")
    abs_err_h = np.asarray(
        [float(row["energy_abs_err_ev"]) / HARTREE_TO_EV for row in pred_rows],
        dtype=np.float64,
    )
    history_row: dict[str, float] = {
        "step": 0,
        "loss": float(np.mean(np.square(abs_err_h))),
        "energy_mae_h": float(np.mean(abs_err_h)),
        "density_mse": float("nan"),
        "density_penalty": float("nan"),
        "density_matrix_mse": float("nan"),
        "density_matrix_penalty": float("nan"),
        "grad_norm": float("nan"),
        "grad_abs_max": float("nan"),
        "param_update_norm": float("nan"),
        "nonfinite_grad_fraction": float("nan"),
        "lr": float("nan"),
    }
    for row, err_h in zip(pred_rows, abs_err_h):
        label = _r_column_label(float(row["r_angstrom"]))
        predicted_energy_h = float(row["predicted_energy_h"])
        reference_energy_h = float(row["reference_energy_h"])
        history_row[f"loss_{label}"] = float(err_h * err_h)
        history_row[f"energy_mae_h_{label}"] = float(err_h)
        history_row[f"energy_signed_err_h_{label}"] = predicted_energy_h - reference_energy_h
        history_row[f"predicted_energy_h_{label}"] = predicted_energy_h
        history_row[f"density_mse_{label}"] = float("nan")
        history_row[f"density_matrix_mse_{label}"] = float("nan")
        history_row[f"grad_norm_{label}"] = float("nan")
        history_row[f"grad_abs_max_{label}"] = float("nan")
        history_row[f"nonfinite_grad_fraction_{label}"] = float("nan")
        history_row[f"scf_converged_{label}"] = float("nan")
        history_row[f"scf_cycles_{label}"] = float("nan")
        history_row[f"scf_selected_rms_density_{label}"] = float("nan")
        history_row[f"scf_best_rms_density_{label}"] = float("nan")
    return [history_row]


def reference_point_csv_row(point: ReferencePoint) -> dict[str, float | str]:
    return {
        "r_angstrom": float(point.r_angstrom),
        "reference_energy_h": float(point.reference_energy_h),
        "mean_field_energy_h": float(point.mean_field_energy_h),
        "correlation_energy_h": float(point.correlation_energy_h),
        "perturbative_corr_h": float(point.perturbative_corr_h),
        "reference_method": str(point.reference_method),
        "n_grid": int(np.asarray(point.molecule.grid.weights).size),
        "n_ao": int(np.asarray(point.molecule.mo_coeff).shape[-1]),
        "electron_count": float(np.asarray(point.molecule.mo_occ).sum()),
    }


def write_reference_points_csv(path: Path, points: list[ReferencePoint]) -> None:
    rows = [reference_point_csv_row(point) for point in points]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _spin_summed_rdm1(molecule: Any) -> jnp.ndarray:
    rdm1 = jnp.asarray(molecule.rdm1)
    return rdm1.sum(axis=0) if rdm1.ndim == 3 else rdm1


def _prediction_scf_initial_molecule(molecule: Any, init_mode: str) -> Any:
    mode = str(init_mode)
    if mode == "cached":
        return molecule
    if mode == "none":
        return replace(molecule, scf_initial_density=None)
    if mode == "rdm1":
        return replace(molecule, scf_initial_density=_spin_summed_rdm1(molecule))
    raise ValueError(f"Unsupported prediction SCF init mode: {init_mode!r}")


def _info_scalar(info: Any, name: str) -> float:
    return float(np.asarray(jax.device_get(jnp.asarray(getattr(info, name)))))


def evaluate_prediction_scf_stability(
    *,
    params: Any,
    functional: Any,
    point: ReferencePoint,
    training_config: GroundStateTrainingConfig,
    init_modes: tuple[str, ...],
    selections: tuple[str, ...],
    tol_ev: float,
    repeats: int = 1,
) -> dict[str, float]:
    energies_h: list[float] = []
    converged_flags: list[bool] = []
    cycles: list[float] = []
    selected_rms: list[float] = []
    for init_mode in init_modes:
        molecule = _prediction_scf_initial_molecule(point.molecule, init_mode)
        for selection in selections:
            cfg = replace(training_config, scf_iterate_selection=str(selection))
            for _repeat in range(max(1, int(repeats))):
                scf_molecule, info = _resolve_training_molecule_and_info_with_mode(
                    params,
                    functional,
                    molecule,
                    cfg,
                )
                energy_h = float(
                    _predict_ground_state_total_energy_from_molecule(
                        params,
                        functional,
                        scf_molecule,
                    )
                )
                energies_h.append(energy_h)
                converged_flags.append(bool(np.asarray(jax.device_get(info.converged))))
                cycles.append(_info_scalar(info, "cycles"))
                selected_rms.append(_info_scalar(info, "selected_rms_density"))
    if not energies_h:
        raise ValueError("At least one SCF stability probe is required.")
    min_energy_h = min(energies_h)
    max_energy_h = max(energies_h)
    spread_ev = abs(max_energy_h - min_energy_h) * HARTREE_TO_EV
    converged_all = all(converged_flags)
    stable = bool(converged_all and spread_ev <= float(tol_ev))
    return {
        "scf_stable": float(int(stable)),
        "scf_energy_spread_ev": float(spread_ev),
        "scf_min_energy_h": float(min_energy_h),
        "scf_max_energy_h": float(max_energy_h),
        "scf_converged_all": float(int(converged_all)),
        "scf_max_cycles": float(max(cycles)),
        "scf_max_selected_rms_density": float(max(selected_rms)),
    }


def make_prediction_stability_evaluator(
    args: argparse.Namespace,
    logger: RunLogger,
) -> Any | None:
    if not bool(args.prediction_scf_stability_check):
        return None
    logger.log(
        "[prediction_scf] stability check enabled "
        f"tol_ev={float(args.prediction_scf_stability_tol_ev):.6g} "
        f"init_modes={list(args.prediction_scf_stability_init_modes)} "
        f"selections={list(args.prediction_scf_stability_selections)} "
        f"repeats={int(args.prediction_scf_stability_repeats)}"
    )

    def stability_evaluator(**kwargs: Any) -> dict[str, float]:
        result = evaluate_prediction_scf_stability(
            **kwargs,
            init_modes=tuple(str(value) for value in args.prediction_scf_stability_init_modes),
            selections=tuple(str(value) for value in args.prediction_scf_stability_selections),
            tol_ev=float(args.prediction_scf_stability_tol_ev),
            repeats=int(args.prediction_scf_stability_repeats),
        )
        point = kwargs["point"]
        logger.log(
            "[prediction_scf] "
            f"R={float(point.r_angstrom):.4f} A "
            f"stable={int(result['scf_stable'])} "
            f"spread={result['scf_energy_spread_ev']:.6e} eV "
            f"converged_all={int(result['scf_converged_all'])} "
            f"max_cycles={result['scf_max_cycles']:.0f} "
            f"max_rms={result['scf_max_selected_rms_density']:.3e}"
        )
        return result

    return stability_evaluator


def summarize_prediction_scf(
    args: argparse.Namespace,
    pred_rows: list[dict[str, float]],
) -> tuple[bool | None, float | None, list[float]]:
    if not bool(args.prediction_scf_stability_check):
        return None, None, []
    prediction_scf_all_stable = all(
        float(row.get("scf_stable", 0.0)) >= 0.5 for row in pred_rows
    )
    prediction_scf_max_energy_spread_ev = float(
        max(float(row.get("scf_energy_spread_ev", 0.0)) for row in pred_rows)
    )
    unstable_prediction_scf_r_values = [
        float(row["r_angstrom"])
        for row in pred_rows
        if float(row.get("scf_stable", 0.0)) < 0.5
    ]
    return (
        prediction_scf_all_stable,
        prediction_scf_max_energy_spread_ev,
        unstable_prediction_scf_r_values,
    )


def prediction_csv_row(
    point: ReferencePoint,
    *,
    params: Any,
    functional: Any,
    training_config: GroundStateTrainingConfig,
    stability_evaluator: Any | None = None,
) -> dict[str, float | str]:
    pred = float(
        predict_ground_state_total_energy(
            params,
            functional,
            point.molecule,
            training_config=training_config,
        )
    )
    row: dict[str, float | str] = {
        "r_angstrom": float(point.r_angstrom),
        "reference_energy_h": float(point.reference_energy_h),
        "predicted_energy_h": pred,
        "energy_abs_err_ev": abs(pred - float(point.reference_energy_h)) * HARTREE_TO_EV,
        "mean_field_energy_h": float(point.mean_field_energy_h),
        "correlation_energy_h": float(point.correlation_energy_h),
        "perturbative_corr_h": float(point.perturbative_corr_h),
        "reference_method": str(point.reference_method),
    }
    if stability_evaluator is not None:
        row.update(
            {
                key: float(value)
                for key, value in stability_evaluator(
                    params=params,
                    functional=functional,
                    point=point,
                    training_config=training_config,
                ).items()
            }
        )
    return row


def write_prediction_csv(
    path: Path,
    points: list[ReferencePoint],
    *,
    params: Any,
    functional: Any,
    training_config: GroundStateTrainingConfig,
    stability_evaluator: Any | None = None,
) -> list[dict[str, float | str]]:
    rows = [
        prediction_csv_row(
            point,
            params=params,
            functional=functional,
            training_config=training_config,
            stability_evaluator=stability_evaluator,
        )
        for point in points
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return rows


def plot_outputs(
    outdir: Path,
    history: list[dict[str, float]],
    pred_rows: list[dict[str, float]],
    *,
    reference_label: str,
    bond_label: str = "N-N",
) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    steps = np.asarray([row["step"] for row in history], dtype=np.float64)
    loss = np.asarray([row["loss"] for row in history], dtype=np.float64)
    mae_ev = np.asarray([row["energy_mae_h"] for row in history], dtype=np.float64) * HARTREE_TO_EV
    fig, axes = plt.subplots(1, 2, figsize=(10.5, 3.8))
    axes[0].plot(steps, np.maximum(loss, 1e-18))
    axes[0].set_yscale("log")
    axes[0].set_xlabel("epoch")
    axes[0].set_ylabel("loss")
    axes[0].grid(alpha=0.25)
    axes[1].plot(steps, np.maximum(mae_ev, 1e-18))
    axes[1].set_yscale("log")
    axes[1].set_xlabel("epoch")
    axes[1].set_ylabel("energy MAE (eV)")
    axes[1].grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(outdir / "training_loss.png", dpi=220)
    plt.close(fig)

    r = np.asarray([row["r_angstrom"] for row in pred_rows], dtype=np.float64)
    target = np.asarray([row["reference_energy_h"] for row in pred_rows], dtype=np.float64)
    pred = np.asarray([row["predicted_energy_h"] for row in pred_rows], dtype=np.float64)
    err = np.asarray([row["energy_abs_err_ev"] for row in pred_rows], dtype=np.float64)
    fig, axes = plt.subplots(1, 2, figsize=(10.5, 3.8))
    axes[0].plot(r, target, "o-", label=reference_label)
    axes[0].plot(r, pred, "s-", label="Neural XC")
    axes[0].set_xlabel(f"{bond_label} distance (Angstrom)")
    axes[0].set_ylabel("total energy (Hartree)")
    axes[0].grid(alpha=0.25)
    axes[0].legend(frameon=False)
    axes[1].plot(r, np.maximum(err, 1e-16), "o-")
    axes[1].set_yscale("log")
    axes[1].set_xlabel(f"{bond_label} distance (Angstrom)")
    axes[1].set_ylabel("absolute error (eV)")
    axes[1].grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(outdir / "n2_ccsdt_ground_curve_train_points.png", dpi=220)
    plt.close(fig)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Train Neural XC on N2 ground-state dissociation points.")
    p.add_argument("--atom-symbol", default="N")
    p.add_argument("--basis", default="def2-svp")
    p.add_argument("--xc", default="b3lyp")
    p.add_argument("--r-min", type=float, default=0.05)
    p.add_argument("--r-max", type=float, default=5.0)
    p.add_argument("--train-points", type=int, default=5)
    p.add_argument("--r-values", type=float, nargs="+", default=None)
    p.add_argument("--steps", type=int, default=1000)
    p.add_argument(
        "--eval-only-checkpoint",
        default=None,
        help="Optional Flax msgpack params checkpoint to evaluate without training.",
    )
    p.add_argument("--reference-cache", default=None)
    p.add_argument("--rebuild-reference-cache", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument("--stream-train", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument("--learning-rate", type=float, default=1e-3)
    p.add_argument("--lr-decay-every", type=int, default=200)
    p.add_argument("--lr-decay-factor", type=float, default=0.5)
    p.add_argument("--seed", type=int, default=0)
    p.add_argument("--hidden-dims", type=int, nargs="+", default=list(DEFAULT_NETWORK_HIDDEN_DIMS))
    p.add_argument(
        "--network-architecture",
        choices=("simple_mlp", "graddft_residual"),
        default=DEFAULT_NETWORK_ARCHITECTURE,
    )
    p.add_argument(
        "--input-feature-mode",
        choices=("enhanced", "canonical", "dm21_original"),
        default=DEFAULT_INPUT_FEATURE_MODE,
    )
    p.add_argument("--include-pt2-channel", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument(
        "--include-hfx-channel",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Add a projected local exact-exchange channel to the Neural_xc basis.",
    )
    p.add_argument(
        "--pt2-channel-mode",
        choices=("scaled_projected", "local_exact"),
        default="local_exact",
    )
    p.add_argument("--allow-experimental-jax-xc", action="store_true")
    p.add_argument("--semilocal-xc", nargs="+", default=list(_DEFAULT_SEMILOCAL_XC))
    p.add_argument("--energy-mse-weight", type=float, default=1.0)
    p.add_argument("--energy-mae-weight", type=float, default=0.0)
    p.add_argument(
        "--energy-normalization",
        choices=("none", "per_electron", "per_atom"),
        default="none",
    )
    p.add_argument(
        "--training-mode",
        choices=("fixed_density", "self_consistent"),
        default="self_consistent",
    )
    p.add_argument("--density-constraint-weight", type=float, default=0.0)
    p.add_argument("--density-matrix-constraint-weight", type=float, default=0.0)
    p.add_argument("--coefficient-prior-weight", type=float, default=0.0)
    p.add_argument("--grids-level", type=int, default=2)
    p.add_argument("--reference-scf-max-cycle", type=int, default=160)
    p.add_argument("--reference-scf-conv-tol", type=float, default=1e-10)
    p.add_argument("--reference-scf-device", choices=("cpu", "gpu"), default="cpu")
    p.add_argument("--jk-backend", choices=("full", "df"), default="full")
    p.add_argument("--hfx-nu-storage", choices=("auto", "dense", "chunked"), default="auto")
    p.add_argument("--ground-state-hf-mode", choices=("off", "nograd", "scf"), default="off")
    p.add_argument("--ground-state-pt2-mode", choices=("off", "nograd", "scf"), default="off")
    p.add_argument(
        "--reference-method",
        choices=("ccsd_t", "casscf", "casscf_nevpt2", "graddft_data"),
        default="ccsd_t",
    )
    p.add_argument(
        "--graddft-dissociation-xlsx",
        "--graddft-dissociation-file",
        dest="graddft_dissociation_xlsx",
        default="data/raw/dissociation/N2_dissociation.xlsx",
    )
    p.add_argument("--active-space", choices=("canonical", "avas"), default="avas")
    p.add_argument("--active-labels", nargs="+", default=["N 2s", "N 2p"])
    p.add_argument("--ncas", type=int, default=8)
    p.add_argument("--nelecas", type=int, default=10)
    p.add_argument("--avas-threshold", type=float, default=0.1)
    p.add_argument("--rhf-max-cycle", type=int, default=200)
    p.add_argument("--rhf-conv-tol", type=float, default=1e-10)
    p.add_argument("--ccsd-max-cycle", type=int, default=120)
    p.add_argument("--ccsd-conv-tol", type=float, default=1e-8)
    p.add_argument("--casscf-max-cycle", type=int, default=80)
    p.add_argument("--casscf-conv-tol", type=float, default=1e-8)
    p.add_argument("--train-scf-max-cycle", type=int, default=0)
    p.add_argument("--train-scf-damping", type=float, default=0.25)
    p.add_argument("--train-scf-conv-tol-energy", type=float, default=1e-6)
    p.add_argument(
        "--train-scf-convergence-metric",
        choices=("energy_and_residual", "energy"),
        default="energy",
    )
    p.add_argument("--train-scf-conv-tol-density", type=float, default=1e-8)
    p.add_argument("--train-scf-vxc-clip", type=float, default=20.0)
    p.add_argument(
        "--scf-iterate-selection",
        choices=("final", "best_rms", "first_converged"),
        default="best_rms",
    )
    p.add_argument(
        "--scf-gradient-mode",
        choices=("expl", "impl"),
        default="impl",
    )
    p.add_argument("--scf-implicit-diff-tolerance", type=float, default=1e-6)
    p.add_argument("--scf-implicit-diff-regularization", type=float, default=1e-3)
    p.add_argument("--scf-require-convergence", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument("--prediction-scf-stability-check", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument("--prediction-scf-stability-tol-ev", type=float, default=0.02)
    p.add_argument(
        "--prediction-scf-stability-init-modes",
        nargs="+",
        choices=("cached", "rdm1", "none"),
        default=["cached", "rdm1"],
    )
    p.add_argument(
        "--prediction-scf-stability-selections",
        nargs="+",
        choices=("final", "best_rms", "first_converged"),
        default=["best_rms", "final"],
    )
    p.add_argument("--prediction-scf-stability-repeats", type=int, default=1)
    p.add_argument("--fail-on-unstable-prediction-scf", action=argparse.BooleanOptionalAction, default=False)
    p.add_argument("--jit-train", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--jit-eval", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--log-every", type=int, default=10)
    p.add_argument("--outdir", default="outputs/n2_ccsdt_ground_train5")
    return p.parse_args(argv)


def run_eval_only_checkpoint_streaming(
    *,
    args: argparse.Namespace,
    logger: RunLogger,
    outdir: Path,
    r_values: np.ndarray,
    graddft_targets: dict[float, float],
    cache_path: Path | None,
    reference_label: str,
    uses_active_space: bool,
) -> dict[str, Any]:
    functional, training_config = build_functional_and_training_config(args)
    stability_evaluator = make_prediction_stability_evaluator(args, logger)
    params = None
    pred_rows: list[dict[str, float | str]] = []
    reference_csv = outdir / "n2_ccsdt_reference_points.csv"
    prediction_csv = outdir / "n2_ccsdt_ground_predictions.csv"
    t0 = time.perf_counter()
    ref_writer = None
    pred_writer = None
    with reference_csv.open("w", newline="", encoding="utf-8") as ref_handle, prediction_csv.open(
        "w",
        newline="",
        encoding="utf-8",
    ) as pred_handle:
        for idx, r_value in enumerate(r_values, start=1):
            point, source = load_or_build_reference_point(
                float(r_value),
                args=args,
                cache_path=cache_path,
                graddft_targets=graddft_targets,
            )
            logger.log(
                f"[ref] {idx:3d}/{len(r_values):3d} {source} "
                f"R={point.r_angstrom:.4f} A "
                f"E_ref={point.reference_energy_h:.10f} Eh "
                f"E_MF={point.mean_field_energy_h:.10f} Eh "
                f"E_pert={point.perturbative_corr_h:.10f} Eh "
                f"grid_n={int(np.asarray(point.molecule.grid.weights).size)}"
            )
            ref_row = reference_point_csv_row(point)
            if ref_writer is None:
                ref_writer = csv.DictWriter(ref_handle, fieldnames=list(ref_row.keys()))
                ref_writer.writeheader()
            ref_writer.writerow(ref_row)
            ref_handle.flush()

            if params is None:
                base_optimizer = optax.adam(float(args.learning_rate))
                template_state = create_train_state_from_molecule(
                    functional,
                    jax.random.PRNGKey(int(args.seed)),
                    point.molecule,
                    base_optimizer,
                )
                params = load_params_checkpoint(
                    args.eval_only_checkpoint,
                    template=template_state.params,
                )
                logger.log(
                    "[eval] loaded checkpoint "
                    f"{args.eval_only_checkpoint} param_norm={_tree_l2_norm(params):.8e}"
                )

            pred_row = prediction_csv_row(
                point,
                params=params,
                functional=functional,
                training_config=training_config,
                stability_evaluator=stability_evaluator,
            )
            if pred_writer is None:
                pred_writer = csv.DictWriter(pred_handle, fieldnames=list(pred_row.keys()))
                pred_writer.writeheader()
            pred_writer.writerow(pred_row)
            pred_handle.flush()
            pred_rows.append(pred_row)
            logger.log(
                "[eval] "
                f"{idx:3d}/{len(r_values):3d} "
                f"R={float(pred_row['r_angstrom']):.4f} A "
                f"abs_err={float(pred_row['energy_abs_err_ev']):.8e} eV"
            )
            del point
            gc.collect()

    if not pred_rows:
        raise ValueError("No prediction rows were produced.")
    (
        prediction_scf_all_stable,
        prediction_scf_max_energy_spread_ev,
        unstable_prediction_scf_r_values,
    ) = summarize_prediction_scf(args, pred_rows)
    history_csv = outdir / "training_history.csv"
    per_point_history_csv = outdir / "training_per_point_history.csv"
    eval_history = prediction_rows_to_eval_history(pred_rows)
    write_history_csv(history_csv, eval_history)
    eval_points = [
        ReferencePoint(
            r_angstrom=float(row["r_angstrom"]),
            molecule=None,
            reference_energy_h=float(row["reference_energy_h"]),
            mean_field_energy_h=float(row["mean_field_energy_h"]),
            correlation_energy_h=float(row["correlation_energy_h"]),
            perturbative_corr_h=float(row["perturbative_corr_h"]),
            reference_method=str(row["reference_method"]),
            target_density_matrix=None,
        )
        for row in pred_rows
    ]
    write_per_point_history_csv(per_point_history_csv, eval_history, eval_points)
    try:
        plot_outputs(
            outdir,
            eval_history,
            pred_rows,
            reference_label=reference_label,
            bond_label=f"{args.atom_symbol}-{args.atom_symbol}",
        )
    except Exception as exc:
        logger.log(f"[plot] skipped after error: {exc!r}")
    summary = {
        "molecule": f"{args.atom_symbol}2",
        "basis": str(args.basis),
        "jk_backend": str(args.jk_backend),
        "hfx_nu_storage": str(args.hfx_nu_storage),
        "reference": reference_label,
        "reference_method": str(args.reference_method),
        "graddft_dissociation_xlsx": str(args.graddft_dissociation_xlsx)
        if str(args.reference_method) == "graddft_data"
        else None,
        "active_space": str(args.active_space) if uses_active_space else None,
        "active_labels": list(args.active_labels) if uses_active_space else None,
        "ncas": int(args.ncas) if uses_active_space else None,
        "nelecas": int(args.nelecas) if uses_active_space else None,
        "eval_only": True,
        "eval_only_checkpoint": str(args.eval_only_checkpoint),
        "include_pt2_channel": bool(args.include_pt2_channel),
        "include_hfx_channel": bool(args.include_hfx_channel),
        "ground_state_hf_mode": str(args.ground_state_hf_mode),
        "ground_state_pt2_mode": str(args.ground_state_pt2_mode),
        "pt2_channel_mode": str(args.pt2_channel_mode) if bool(args.include_pt2_channel) else None,
        "density_constraint_weight": float(args.density_constraint_weight),
        "density_matrix_constraint_weight": float(args.density_matrix_constraint_weight),
        "reference_scf_device": str(args.reference_scf_device),
        "steps": 0,
        "learning_rate": float(args.learning_rate),
        "lr_decay_every": int(args.lr_decay_every),
        "lr_decay_factor": float(args.lr_decay_factor),
        "train_r_values_angstrom": [float(value) for value in r_values],
        "elapsed_s": float(time.perf_counter() - t0),
        "final_loss": float(eval_history[0]["loss"]),
        "final_energy_mae_ev": float(eval_history[0]["energy_mae_h"]) * HARTREE_TO_EV,
        "min_loss": float(eval_history[0]["loss"]),
        "min_loss_step": 0,
        "prediction_energy_mae_ev": float(np.mean([float(row["energy_abs_err_ev"]) for row in pred_rows])),
        "prediction_scf_stability_check": bool(args.prediction_scf_stability_check),
        "prediction_scf_stability_tol_ev": float(args.prediction_scf_stability_tol_ev),
        "prediction_scf_all_stable": prediction_scf_all_stable,
        "prediction_scf_max_energy_spread_ev": prediction_scf_max_energy_spread_ev,
        "unstable_prediction_scf_r_values_angstrom": unstable_prediction_scf_r_values,
        "reference_points_csv": str(reference_csv),
        "training_history_csv": str(history_csv),
        "training_per_point_history_csv": str(per_point_history_csv),
        "prediction_csv": str(prediction_csv),
        "training_curve_png": str(outdir / "training_loss.png"),
        "prediction_curve_png": str(outdir / "n2_ccsdt_ground_curve_train_points.png"),
        "visualization_manifest": str(outdir / "visualization_manifest.json"),
    }
    (outdir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    prediction_curve_y = [
        "reference_energy_h",
        "predicted_energy_h",
        "energy_abs_err_ev",
        "mean_field_energy_h",
        "correlation_energy_h",
        "perturbative_corr_h",
    ]
    if bool(args.prediction_scf_stability_check):
        prediction_curve_y.extend(_PREDICTION_SCF_STABILITY_FIELDS)
    visualization_manifest = {
        "paper_experiment": "Ground-State Potential-Energy Surfaces",
        "description": f"Checkpoint-only N2 {reference_label} ground-state evaluation visualizations.",
        "figures": [
            {
                "figure": str(outdir / "training_loss.png"),
                "data_files": [str(history_csv), str(per_point_history_csv)],
                "x": "step",
                "y": [
                    "loss",
                    "energy_mae_h",
                    "density_mse",
                    "density_penalty",
                    "density_matrix_mse",
                    "density_matrix_penalty",
                ],
            },
            {
                "figure": str(outdir / "n2_ccsdt_ground_curve_train_points.png"),
                "data_files": [str(prediction_csv), str(reference_csv)],
                "x": "r_angstrom",
                "y": prediction_curve_y,
            },
        ],
        "metadata_files": [str(outdir / "summary.json")],
    }
    (outdir / "visualization_manifest.json").write_text(
        json.dumps(visualization_manifest, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    logger.log(
        "[summary] "
        f"eval_only=True final_energy_mae={summary['final_energy_mae_ev']:.8e} eV "
        f"prediction_energy_mae={summary['prediction_energy_mae_ev']:.8e} eV "
        f"prediction_scf_all_stable={summary['prediction_scf_all_stable']} "
        f"prediction_scf_max_spread={summary['prediction_scf_max_energy_spread_ev']} "
        f"outdir={outdir}"
    )
    if (
        bool(args.fail_on_unstable_prediction_scf)
        and bool(args.prediction_scf_stability_check)
        and prediction_scf_all_stable is False
    ):
        raise RuntimeError(
            "Unstable prediction SCF detected at R values "
            f"{unstable_prediction_scf_r_values}; max spread="
            f"{prediction_scf_max_energy_spread_ev:.6g} eV"
        )
    return summary


def main(argv: list[str] | None = None) -> dict[str, Any]:
    args = parse_args(argv)
    reference_label = _reference_label(args)
    uses_active_space = str(args.reference_method) in {"casscf", "casscf_nevpt2"}
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    logger = RunLogger(outdir / "run.log")
    logger.log(
        "Config: "
        f"molecule={args.atom_symbol}2 basis={args.basis} xc={args.xc} "
        f"R=[{args.r_min},{args.r_max}] train_points={args.train_points} "
        f"r_values={args.r_values if args.r_values is not None else 'linspace'} "
        f"steps={args.steps} lr={args.learning_rate} "
        f"lr_decay_every={args.lr_decay_every} lr_decay_factor={args.lr_decay_factor} "
        f"jk_backend={args.jk_backend} hfx_nu_storage={args.hfx_nu_storage} "
        f"include_pt2_channel={bool(args.include_pt2_channel)} "
        f"include_hfx_channel={bool(args.include_hfx_channel)} "
        f"ground_state_hf_mode={args.ground_state_hf_mode} "
        f"ground_state_pt2_mode={args.ground_state_pt2_mode} "
        f"pt2_channel_mode={args.pt2_channel_mode if bool(args.include_pt2_channel) else 'none'} "
        f"training_mode={args.training_mode} "
        f"density_constraint_weight={args.density_constraint_weight} "
        f"density_matrix_constraint_weight={args.density_matrix_constraint_weight} "
        f"reference_scf_device={args.reference_scf_device} "
        f"reference={reference_label} "
        f"active_space={args.active_space if uses_active_space else 'none'}"
    )
    r_values = (
        np.asarray([float(value) for value in args.r_values], dtype=np.float64)
        if args.r_values is not None
        else np.linspace(float(args.r_min), float(args.r_max), int(args.train_points))
    )
    graddft_targets = (
        load_graddft_dissociation_targets(
            args.graddft_dissociation_xlsx,
            r_values=r_values,
        )
        if str(args.reference_method) == "graddft_data"
        else {}
    )
    if graddft_targets:
        logger.log(
            f"[ref] loaded GradDFT dissociation targets from {args.graddft_dissociation_xlsx}"
        )
    points: list[ReferencePoint]
    cache_path = Path(str(args.reference_cache)) if args.reference_cache is not None else None
    cache_group = _cache_group_name(args)
    if args.eval_only_checkpoint is not None:
        return run_eval_only_checkpoint_streaming(
            args=args,
            logger=logger,
            outdir=outdir,
            r_values=r_values,
            graddft_targets=graddft_targets,
            cache_path=cache_path,
            reference_label=reference_label,
            uses_active_space=uses_active_space,
        )
    if (
        cache_path is not None
        and _has_hdf5_group(cache_path, cache_group)
        and not bool(args.rebuild_reference_cache)
    ):
        logger.log(f"[ref] loading cached references from {cache_path}:{cache_group}")
        points = _load_reference_points_hdf5(
            cache_path,
            cache_group,
            hfx_nu_storage=str(args.hfx_nu_storage),
        )
        logger.log(f"[ref] loaded {len(points)} cached references")
    else:
        points = []
        t_ref = time.perf_counter()
        for idx, r_value in enumerate(r_values, start=1):
            point_group = _point_cache_group_name(args, float(r_value))
            if (
                cache_path is not None
                and _has_hdf5_group(cache_path, point_group)
                and not bool(args.rebuild_reference_cache)
            ):
                point = _load_reference_point_hdf5(
                    cache_path,
                    point_group,
                    hfx_nu_storage=str(args.hfx_nu_storage),
                )
                source = "cache"
            else:
                point = build_reference_point(
                    float(r_value),
                    args=args,
                    graddft_target_energy_h=graddft_targets.get(float(r_value)),
                )
                source = "built"
                if cache_path is not None:
                    _save_reference_point_hdf5(cache_path, point_group, point)
                    if str(args.hfx_nu_storage) == "chunked":
                        point = _load_reference_point_hdf5(
                            cache_path,
                            point_group,
                            hfx_nu_storage="chunked",
                        )
            points.append(point)
            logger.log(
                f"[ref] {idx:3d}/{len(r_values):3d} {source} "
                f"R={point.r_angstrom:.4f} A "
                f"E_ref={point.reference_energy_h:.10f} Eh "
                f"E_MF={point.mean_field_energy_h:.10f} Eh "
                f"E_pert={point.perturbative_corr_h:.10f} Eh "
                f"grid_n={int(np.asarray(point.molecule.grid.weights).size)}"
            )
        logger.log(f"[ref] done in {time.perf_counter() - t_ref:.2f} s")
        if cache_path is not None:
            logger.log(f"[ref] writing cached references to {cache_path}:{cache_group}")
            _save_reference_points_hdf5(cache_path, cache_group, points)

    reference_csv = outdir / "n2_ccsdt_reference_points.csv"
    write_reference_points_csv(reference_csv, points)
    logger.log(f"Wrote reference csv: {reference_csv}")

    if args.eval_only_checkpoint is not None:
        functional, training_config = build_functional_and_training_config(args)
        base_optimizer = optax.adam(float(args.learning_rate))
        template_state = create_train_state_from_molecule(
            functional,
            jax.random.PRNGKey(int(args.seed)),
            points[0].molecule,
            base_optimizer,
        )
        params = load_params_checkpoint(
            args.eval_only_checkpoint,
            template=template_state.params,
        )
        logger.log(
            "[eval] loaded checkpoint "
            f"{args.eval_only_checkpoint} param_norm={_tree_l2_norm(params):.8e}"
        )
        stability_evaluator = make_prediction_stability_evaluator(args, logger)
        pred_rows = write_prediction_csv(
            outdir / "n2_ccsdt_ground_predictions.csv",
            points,
            params=params,
            functional=functional,
            training_config=training_config,
            stability_evaluator=stability_evaluator,
        )
        (
            prediction_scf_all_stable,
            prediction_scf_max_energy_spread_ev,
            unstable_prediction_scf_r_values,
        ) = summarize_prediction_scf(args, pred_rows)
        history_csv = outdir / "training_history.csv"
        per_point_history_csv = outdir / "training_per_point_history.csv"
        eval_history = prediction_rows_to_eval_history(pred_rows)
        write_history_csv(history_csv, eval_history)
        write_per_point_history_csv(per_point_history_csv, eval_history, points)
        try:
            plot_outputs(
                outdir,
                eval_history,
                pred_rows,
                reference_label=reference_label,
                bond_label=f"{args.atom_symbol}-{args.atom_symbol}",
            )
        except Exception as exc:
            logger.log(f"[plot] skipped after error: {exc!r}")
        summary = {
            "molecule": f"{args.atom_symbol}2",
            "basis": str(args.basis),
            "jk_backend": str(args.jk_backend),
            "hfx_nu_storage": str(args.hfx_nu_storage),
            "reference": reference_label,
            "reference_method": str(args.reference_method),
            "graddft_dissociation_xlsx": str(args.graddft_dissociation_xlsx)
            if str(args.reference_method) == "graddft_data"
            else None,
            "active_space": str(args.active_space) if uses_active_space else None,
            "active_labels": list(args.active_labels) if uses_active_space else None,
            "ncas": int(args.ncas) if uses_active_space else None,
            "nelecas": int(args.nelecas) if uses_active_space else None,
            "eval_only": True,
            "eval_only_checkpoint": str(args.eval_only_checkpoint),
            "include_pt2_channel": bool(args.include_pt2_channel),
            "include_hfx_channel": bool(args.include_hfx_channel),
            "ground_state_hf_mode": str(args.ground_state_hf_mode),
            "ground_state_pt2_mode": str(args.ground_state_pt2_mode),
            "pt2_channel_mode": str(args.pt2_channel_mode) if bool(args.include_pt2_channel) else None,
            "density_constraint_weight": float(args.density_constraint_weight),
            "density_matrix_constraint_weight": float(args.density_matrix_constraint_weight),
            "reference_scf_device": str(args.reference_scf_device),
            "steps": 0,
            "learning_rate": float(args.learning_rate),
            "lr_decay_every": int(args.lr_decay_every),
            "lr_decay_factor": float(args.lr_decay_factor),
            "train_r_values_angstrom": [float(value) for value in r_values],
            "final_loss": float(eval_history[0]["loss"]),
            "final_energy_mae_ev": float(eval_history[0]["energy_mae_h"]) * HARTREE_TO_EV,
            "min_loss": float(eval_history[0]["loss"]),
            "min_loss_step": 0,
            "prediction_energy_mae_ev": float(np.mean([row["energy_abs_err_ev"] for row in pred_rows])),
            "prediction_scf_stability_check": bool(args.prediction_scf_stability_check),
            "prediction_scf_stability_tol_ev": float(args.prediction_scf_stability_tol_ev),
            "prediction_scf_all_stable": prediction_scf_all_stable,
            "prediction_scf_max_energy_spread_ev": prediction_scf_max_energy_spread_ev,
            "unstable_prediction_scf_r_values_angstrom": unstable_prediction_scf_r_values,
            "reference_points_csv": str(outdir / "n2_ccsdt_reference_points.csv"),
            "training_history_csv": str(history_csv),
            "training_per_point_history_csv": str(per_point_history_csv),
            "prediction_csv": str(outdir / "n2_ccsdt_ground_predictions.csv"),
            "training_curve_png": str(outdir / "training_loss.png"),
            "prediction_curve_png": str(outdir / "n2_ccsdt_ground_curve_train_points.png"),
            "visualization_manifest": str(outdir / "visualization_manifest.json"),
        }
        (outdir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
        prediction_curve_y = [
            "reference_energy_h",
            "predicted_energy_h",
            "energy_abs_err_ev",
            "mean_field_energy_h",
            "correlation_energy_h",
            "perturbative_corr_h",
        ]
        if bool(args.prediction_scf_stability_check):
            prediction_curve_y.extend(_PREDICTION_SCF_STABILITY_FIELDS)
        visualization_manifest = {
            "paper_experiment": "Ground-State Potential-Energy Surfaces",
            "description": f"Checkpoint-only N2 {reference_label} ground-state evaluation visualizations.",
            "figures": [
                {
                    "figure": str(outdir / "training_loss.png"),
                    "data_files": [str(history_csv), str(per_point_history_csv)],
                    "x": "step",
                    "y": [
                        "loss",
                        "energy_mae_h",
                        "density_mse",
                        "density_penalty",
                        "density_matrix_mse",
                        "density_matrix_penalty",
                    ],
                },
                {
                    "figure": str(outdir / "n2_ccsdt_ground_curve_train_points.png"),
                    "data_files": [
                        str(outdir / "n2_ccsdt_ground_predictions.csv"),
                        str(outdir / "n2_ccsdt_reference_points.csv"),
                    ],
                    "x": "r_angstrom",
                    "y": prediction_curve_y,
                },
            ],
            "metadata_files": [str(outdir / "summary.json")],
        }
        (outdir / "visualization_manifest.json").write_text(
            json.dumps(visualization_manifest, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        logger.log(
            "[summary] "
            f"eval_only=True final_energy_mae={summary['final_energy_mae_ev']:.8e} eV "
            f"prediction_energy_mae={summary['prediction_energy_mae_ev']:.8e} eV "
            f"prediction_scf_all_stable={summary['prediction_scf_all_stable']} "
            f"prediction_scf_max_spread={summary['prediction_scf_max_energy_spread_ev']} "
            f"outdir={outdir}"
        )
        if (
            bool(args.fail_on_unstable_prediction_scf)
            and bool(args.prediction_scf_stability_check)
            and prediction_scf_all_stable is False
        ):
            raise RuntimeError(
                "Unstable prediction SCF detected at R values "
                f"{unstable_prediction_scf_r_values}; max spread="
                f"{prediction_scf_max_energy_spread_ev:.6g} eV"
            )
        return summary

    training = train_functional(points, args=args, logger=logger)
    history_csv = outdir / "training_history.csv"
    per_point_history_csv = outdir / "training_per_point_history.csv"
    write_history_csv(history_csv, training["history"])
    write_per_point_history_csv(per_point_history_csv, training["history"], points)
    stability_evaluator = make_prediction_stability_evaluator(args, logger)
    pred_rows = write_prediction_csv(
        outdir / "n2_ccsdt_ground_predictions.csv",
        points,
        params=training["best_params"],
        functional=training["functional"],
        training_config=training["training_config"],
        stability_evaluator=stability_evaluator,
    )
    (
        prediction_scf_all_stable,
        prediction_scf_max_energy_spread_ev,
        unstable_prediction_scf_r_values,
    ) = summarize_prediction_scf(args, pred_rows)
    save_params_checkpoint(
        outdir / "neural_xc_params.msgpack",
        training["best_params"],
        metadata={
            "molecule": f"{args.atom_symbol}2",
            "basis": str(args.basis),
            "jk_backend": str(args.jk_backend),
            "hfx_nu_storage": str(args.hfx_nu_storage),
            "reference": reference_label,
            "reference_method": str(args.reference_method),
            "graddft_dissociation_xlsx": str(args.graddft_dissociation_xlsx)
            if str(args.reference_method) == "graddft_data"
            else None,
            "active_space": str(args.active_space) if uses_active_space else None,
            "active_labels": list(args.active_labels) if uses_active_space else None,
            "ncas": int(args.ncas) if uses_active_space else None,
            "nelecas": int(args.nelecas) if uses_active_space else None,
            "train_r_values_angstrom": [float(value) for value in r_values],
            "steps": int(args.steps),
            "training_mode": str(args.training_mode),
            "learning_rate": float(args.learning_rate),
            "lr_decay_every": int(args.lr_decay_every),
            "lr_decay_factor": float(args.lr_decay_factor),
            "include_pt2_channel": bool(args.include_pt2_channel),
            "include_hfx_channel": bool(args.include_hfx_channel),
            "ground_state_hf_mode": str(args.ground_state_hf_mode),
            "ground_state_pt2_mode": str(args.ground_state_pt2_mode),
            "pt2_channel_mode": str(args.pt2_channel_mode) if bool(args.include_pt2_channel) else None,
            "density_constraint_weight": float(args.density_constraint_weight),
            "density_matrix_constraint_weight": float(args.density_matrix_constraint_weight),
            "reference_scf_device": str(args.reference_scf_device),
            "prediction_scf_stability_check": bool(args.prediction_scf_stability_check),
            "prediction_scf_stability_tol_ev": float(args.prediction_scf_stability_tol_ev),
            "prediction_scf_all_stable": prediction_scf_all_stable,
            "prediction_scf_max_energy_spread_ev": prediction_scf_max_energy_spread_ev,
            "unstable_prediction_scf_r_values_angstrom": unstable_prediction_scf_r_values,
        },
    )
    try:
        plot_outputs(
            outdir,
            training["history"],
            pred_rows,
            reference_label=reference_label,
            bond_label=f"{args.atom_symbol}-{args.atom_symbol}",
        )
    except Exception as exc:
        logger.log(f"[plot] skipped after error: {exc!r}")
    summary = {
        "molecule": f"{args.atom_symbol}2",
        "basis": str(args.basis),
        "jk_backend": str(args.jk_backend),
        "hfx_nu_storage": str(args.hfx_nu_storage),
        "reference": reference_label,
        "reference_method": str(args.reference_method),
        "graddft_dissociation_xlsx": str(args.graddft_dissociation_xlsx)
        if str(args.reference_method) == "graddft_data"
        else None,
        "active_space": str(args.active_space) if uses_active_space else None,
        "active_labels": list(args.active_labels) if uses_active_space else None,
        "ncas": int(args.ncas) if uses_active_space else None,
        "nelecas": int(args.nelecas) if uses_active_space else None,
        "include_pt2_channel": bool(args.include_pt2_channel),
        "include_hfx_channel": bool(args.include_hfx_channel),
        "ground_state_hf_mode": str(args.ground_state_hf_mode),
        "ground_state_pt2_mode": str(args.ground_state_pt2_mode),
        "pt2_channel_mode": str(args.pt2_channel_mode) if bool(args.include_pt2_channel) else None,
        "training_mode": str(args.training_mode),
        "density_constraint_weight": float(args.density_constraint_weight),
        "density_matrix_constraint_weight": float(args.density_matrix_constraint_weight),
        "reference_scf_device": str(args.reference_scf_device),
        "steps": int(args.steps),
        "learning_rate": float(args.learning_rate),
        "lr_decay_every": int(args.lr_decay_every),
        "lr_decay_factor": float(args.lr_decay_factor),
        "train_r_values_angstrom": [float(value) for value in r_values],
        "elapsed_s": float(training["elapsed_s"]),
        "final_loss": float(training["final_loss"]),
        "final_energy_mae_ev": float(training["final_energy_mae_h"]) * HARTREE_TO_EV,
        "min_loss": float(training["min_loss"]),
        "min_loss_step": int(training["min_loss_step"]),
        "prediction_energy_mae_ev": float(np.mean([row["energy_abs_err_ev"] for row in pred_rows])),
        "prediction_scf_stability_check": bool(args.prediction_scf_stability_check),
        "prediction_scf_stability_tol_ev": float(args.prediction_scf_stability_tol_ev),
        "prediction_scf_all_stable": prediction_scf_all_stable,
        "prediction_scf_max_energy_spread_ev": prediction_scf_max_energy_spread_ev,
        "unstable_prediction_scf_r_values_angstrom": unstable_prediction_scf_r_values,
        "reference_points_csv": str(outdir / "n2_ccsdt_reference_points.csv"),
        "training_history_csv": str(history_csv),
        "training_per_point_history_csv": str(per_point_history_csv),
        "prediction_csv": str(outdir / "n2_ccsdt_ground_predictions.csv"),
        "training_curve_png": str(outdir / "training_loss.png"),
        "prediction_curve_png": str(outdir / "n2_ccsdt_ground_curve_train_points.png"),
        "visualization_manifest": str(outdir / "visualization_manifest.json"),
    }
    (outdir / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    prediction_curve_y = [
        "reference_energy_h",
        "predicted_energy_h",
        "energy_abs_err_ev",
        "mean_field_energy_h",
        "correlation_energy_h",
        "perturbative_corr_h",
    ]
    if bool(args.prediction_scf_stability_check):
        prediction_curve_y.extend(_PREDICTION_SCF_STABILITY_FIELDS)
    visualization_manifest = {
        "paper_experiment": "Ground-State Potential-Energy Surfaces",
        "description": f"Data files needed to reproduce N2 {reference_label} ground-state training visualizations.",
        "figures": [
            {
                "figure": str(outdir / "training_loss.png"),
                "data_files": [str(history_csv), str(per_point_history_csv)],
                "x": "step",
                "y": [
                    "loss",
                    "energy_mae_h",
                    "density_mse",
                    "density_penalty",
                    "density_matrix_mse",
                    "density_matrix_penalty",
                    "grad_norm",
                    "param_update_norm",
                    "scf_converged_fraction",
                    "scf_cycles_mean",
                    "scf_cycles_max",
                    "scf_selected_rms_max",
                ],
            },
            {
                "figure": str(outdir / "n2_ccsdt_ground_curve_train_points.png"),
                "data_files": [
                    str(outdir / "n2_ccsdt_ground_predictions.csv"),
                    str(outdir / "n2_ccsdt_reference_points.csv"),
                ],
                "x": "r_angstrom",
                "y": prediction_curve_y,
            },
        ],
        "metadata_files": [str(outdir / "summary.json"), str(outdir / "neural_xc_params.msgpack.meta.json")],
    }
    (outdir / "visualization_manifest.json").write_text(
        json.dumps(visualization_manifest, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    logger.log(
        "[summary] "
        f"final_energy_mae={summary['final_energy_mae_ev']:.8e} eV "
        f"prediction_energy_mae={summary['prediction_energy_mae_ev']:.8e} eV "
        f"prediction_scf_all_stable={summary['prediction_scf_all_stable']} "
        f"prediction_scf_max_spread={summary['prediction_scf_max_energy_spread_ev']} "
        f"outdir={outdir}"
    )
    if (
        bool(args.fail_on_unstable_prediction_scf)
        and bool(args.prediction_scf_stability_check)
        and prediction_scf_all_stable is False
    ):
        raise RuntimeError(
            "Unstable prediction SCF detected at R values "
            f"{unstable_prediction_scf_r_values}; max spread="
            f"{prediction_scf_max_energy_spread_ev:.6g} eV"
        )
    return summary


if __name__ == "__main__":
    main()
